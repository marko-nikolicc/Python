from tkinter import *
import tkinter.ttk as ttk
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
import os

class PhoneBook():
    def __init__(self):
        self.phonebook = tk.Tk()
        self.phonebook.title("Telefonski imenik")
        self.name = StringVar()
        self.last_name = StringVar()
        self.phone_number = StringVar()
        self.find = StringVar()
        self.find.set("Pretrazi...")
        self.rbutton = IntVar()
        self.list1 = []
        self.list2 = []

        input_data = tk.LabelFrame(self.phonebook, text="Unos podataka", padx=5, pady=5)
        nameLabel = tk.Label(input_data, text="Ime: ")
        last_nameLabel = tk.Label(input_data, text="Prezime: ")
        phone_numberLabel = tk.Label(input_data, text="Telefon: ")
        self.nameEntry = tk.Entry(input_data, textvariable=self.name)
        self.last_nameEntry = tk.Entry(input_data, textvariable=self.last_name)
        self.phone_numberEntry = tk.Entry(input_data, textvariable=self.phone_number)
        Add = tk.Button(input_data, text="Dodaj", command=self.add)

        input_data.grid(row=0, column=0)

        nameLabel.grid(row=0, column=0)
        self.nameEntry.grid(row=0, column=1)
        last_nameLabel.grid(row=1, column=0)
        self.last_nameEntry.grid(row=1, column=1)
        phone_numberLabel.grid(row=2, column=0)
        self.phone_numberEntry.grid(row=2, column=1)
        Add.grid(row=0, column=3, columnspan=3, rowspan=3, sticky=W + E + N + S)

        tools = tk.LabelFrame(self.phonebook, text="Alatke", padx=5, pady=5)
        tools.grid(row=0, column=1)
        clear = tk.Button(tools, text="Obrisi sve", command=self.clearAll)
        clear.grid(row=0, column=0, sticky=W + E)
        edit = tk.Button(tools, text="Izmeni", command=self.edit)
        edit.grid(row=0, column=1, sticky=W + E)
        clearSelected = tk.Button(tools, text="Obrisi selektovano", command=self.clear_selected)
        clearSelected.grid(row=1, column=0, sticky=W + E)
        save = tk.Button(tools, text="Memorisi", command=self.save)
        save.grid(row=1, column=1, sticky=W + E)
        importData = tk.Button(tools, text="Ispisi bazu", command=self.importDatabase)
        importData.grid(row=2, column=1, sticky=W + E)
        deleteDatabase = tk.Button(tools, text="Obrisi bazu", command=self.deleteDatabase)
        deleteDatabase.grid(row=2, column=0, sticky=W + E)

        contact_search = tk.LabelFrame(self.phonebook, text="Pretraga kontakata", padx=5, pady=5)
        contact_search.grid(row=2, column=0, columnspan=2)
        self.searchEntry = tk.Entry(contact_search, textvariable=self.find).pack(side=LEFT)
        searchButton = tk.Button(contact_search, text="Pretrazi", command=self.search).pack(side=LEFT)
        byName = tk.Radiobutton(contact_search, text="Po imenu", variable=self.rbutton, value=1).pack(side=LEFT)
        byLast_name = tk.Radiobutton(contact_search, text="Po prezimenu", variable=self.rbutton, value=2).pack(side=LEFT)
        byPhone_number = tk.Radiobutton(contact_search, text="Po broju telefona", variable=self.rbutton, value=3)
        byPhone_number.pack(side=LEFT)

        self.display = ttk.Treeview(self.phonebook)
        self.display['show'] = 'headings'
        self.display["columns"] = ("name", "lastname", "phone_number")
        self.display.column("name", width=230, minwidth=230)
        self.display.column("lastname", width=230, minwidth=230)
        self.display.column("phone_number", width=140, minwidth=140, stretch=tk.NO)
        self.display.heading("name", text="Ime")
        self.display.heading("lastname", text="Prezime")
        self.display.heading("phone_number", text="Telefon")
        self.display.grid(row=1, column=0, columnspan=2)

        self.nameEntry.focus()
        self.phonebook.mainloop()


    def add(self):
        if self.validation(self.name, self.last_name, self.phone_number, self.nameEntry, self.last_nameEntry, self.phone_numberEntry)==True:
            self.display.insert("", "end", values=((self.name.get()), (self.last_name.get()), (self.phone_number.get())))
            self.list1.append((self.name.get(), self.last_name.get(), self.phone_number.get()))
            self.name.set("")
            self.last_name.set("")
            self.phone_number.set("")
            self.nameEntry.focus()


    def clearAll(self):
        if self.list1 != []:
            self.display.delete(*self.display.get_children())
            self.list1.clear()
        else:
            messagebox.showwarning("Obavestenje","Lista je vec prazna.")


    def validation(self, name, lastname, phone_number, nameEntry, lastnameEntry, phone_numberEntry):
        if name.get()== "":
            messagebox.showerror("Greska", "Ime je obavezan podatak. Unesite ime.")
            nameEntry.focus()
            return False
        elif lastname.get()== "":
            messagebox.showerror("Greska","Prezime je obavezan podatak. Unesite prezime.")
            lastnameEntry.focus()
            return False
        elif phone_number.get()== "":
            messagebox.showerror("Greska","Broj telefona je obavezan podatak. Unesite broj telefona.")
            phone_numberEntry.focus()
            return False
        elif not phone_number.get().isnumeric():
            messagebox.showerror("Greska","Broj telefona moze sadrzati samo numericke karaktere(brojeve).")
            phone_number.set("")
            phone_numberEntry.focus()
            return False
        elif phone_number.get()[:3]!= "381":
            messagebox.showerror("Neispravan broj telefona","Broj telefona mora pocinjati sa 381")
            phone_number.set("")
            phone_numberEntry.focus()
            return False
        elif phone_number.get()[3]== "0":
            messagebox.showerror("Neispravan broj telefona","Broj telefona nije odgovarajuceg formata.")
            phone_number.set("")
            phone_numberEntry.focus()
            return False
        elif len(phone_number.get())<11 or len(phone_number.get())>12:
            messagebox.showerror("Neispravan broj telefona","Neispravna duzina broja telefona.")
            phone_number.set("")
            phone_numberEntry.focus()
            return False
        else:
            return True


    def clear_selected(self):
        if self.list1 != [] and self.display.selection() != ():
            question = messagebox.askquestion("Brisanje","Da li zelite da obrisete ovaj kontakt iz baze?",icon="question")
            if question == "yes":
                confirmation = False
                deleted = 0
                attempts = 0
                databaseAccess = 0
                for selected in self.display.selection():
                    name = str(tuple(self.display.item(selected)['values'])[0])
                    lastname = str(tuple(self.display.item(selected)['values'])[1])
                    phone_number = str(tuple(self.display.item(selected)['values'])[2])
                    for item in range(len(self.list1)):
                        if self.list1[item] == (name, lastname, phone_number):
                            self.list1.pop(item)
                            break
                    try:
                        contacts = load_workbook(filename="ContactsDatabase.xlsx")
                        contacts.active = 0
                        Sheet = contacts.active
                        number = 2
                        for value in Sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
                            if value[0] == name and value[1] == lastname and str(value[2]) == str(phone_number):
                                Sheet.delete_rows(idx=number, amount=1)
                                contacts.save(filename="ContactsDatabase.xlsx")
                                confirmation = True
                                deleted+=1
                            else:
                                number+=1
                    except:
                        if databaseAccess == 0:
                            messagebox.showinfo("Obavestenje", "Ne postoji baza sa kontaktima.")
                            databaseAccess += 1
                    attempts +=1
                    self.display.delete(selected)

                if confirmation == True:
                    if deleted == 1:
                        messagebox.showinfo("Obavestenje", "Oznaceni kontakt je uspesno obrisan iz baze.")
                    elif deleted > 1:
                        messagebox.showinfo("Obavestenje", "Oznaceni kontakti su uspesno obrisani iz baze.")
                else:
                    if attempts == 1:
                        messagebox.showinfo("Obavestenje", "Kontakt koji zelite obrisati ne postoji u bazi kontakata, ali ce biti obrisan sa liste.")
                    elif attempts > 1:
                        messagebox.showinfo("Obavestenje", "Kontakti koje zelite obrisati ne postoje u bazi kontakata, ali ce biti obrisani sa liste.")
            else:
                for selected in self.display.selection():
                    name = str(tuple(self.display.item(selected)['values'])[0])
                    lastname = str(tuple(self.display.item(selected)['values'])[1])
                    phone_number = str(tuple(self.display.item(selected)['values'])[2])
                    for item in range(len(self.list1)):
                        if self.list1[item] == (name, lastname, phone_number):
                            self.list1.pop(item)
                            break
                    self.display.delete(selected)
                messagebox.showinfo("Obavestenje","Kontakt je prividno obrisan sa liste, ali se i dalje nalazi u bazi.")
        else:
            messagebox.showwarning("Obavestenje","Niste oznacili kontakte koje zelite da obrisete.")


    def save(self):
        if self.list1 != []:
            try:
                contacts = load_workbook(filename="ContactsDatabase.xlsx")
            except:
                create = Workbook()
                create.save(filename="ContactsDatabase.xlsx")
                contacts = load_workbook(filename="ContactsDatabase.xlsx")

            contacts.active = 0
            Sheet = contacts.active

            Sheet.column_dimensions['A'].width = 20.0
            Sheet.column_dimensions['B'].width = 20.0
            Sheet.column_dimensions['C'].width = 20.0

            if Sheet["A1"].value == "" or Sheet["A1"].value == None:
                Sheet["A1"] = "NAME"
                Sheet["B1"] = "LASTNAME"
                Sheet["C1"] = "PHONE_NUMBER"

            for contact in self.list1:
                name = contact[0]
                lastname = contact[1]
                phone_number = contact[2]

                number = 2
                for value in Sheet.iter_rows(min_row=1, min_col=1, max_col=3, values_only=True):
                    if Sheet["A" + str(number)].value == "" or Sheet["A" + str(number)].value == None:
                        Sheet["A" + str(number)] = name
                        Sheet["B" + str(number)] = lastname
                        Sheet["C" + str(number)] = phone_number
                        num = 0
                        for value in Sheet.iter_rows(min_row=number+1, min_col=1, max_col=3, values_only=True):
                            if Sheet["A" + str(number+1+num)].value == name and Sheet["B" + str(number+1+num)].value == lastname and Sheet["C" + str(number+1+num)].value == phone_number:
                                Sheet["A" + str(number + 1 + num)] = ""
                                Sheet["B" + str(number + 1 + num)] = ""
                                Sheet["C" + str(number + 1 + num)] = ""
                            num+=1
                    elif Sheet["A" + str(number)].value == name and Sheet["B" + str(number)].value == lastname and Sheet["C" + str(number)].value == phone_number:
                        break
                    else:
                        number+=1

            contacts.save(filename="ContactsDatabase.xlsx")
            messagebox.showinfo("Obavestenje","Uspesno ste snimili kontakte u bazu.")
        else:
            messagebox.showwarning("Obavestenje","Niste dodali kontakte u listu da biste ih memorisali u bazu.")


    def search(self):
        for contact in self.list1:
            if self.rbutton.get()==1 and self.find.get() in contact[0]:
                self.list2.append(contact)
            elif self.rbutton.get()==2 and self.find.get() in contact[1]:
                self.list2.append(contact)
            elif self.rbutton.get()==3 and self.find.get() in contact[2]:
                self.list2.append(contact)

        if self.list2==[]:
            messagebox.showwarning("Obavestenje","Trazeni kontakt nije pronadjen.")
            self.find.set("Pretrazi...")
        else:
            self.display.delete(*self.display.get_children())
            self.list1.clear()
            for contact in self.list2:
                self.display.insert("", "end", values=((contact[0]), (contact[1]), (contact[2])))
                self.list1.append((contact[0], contact[1], contact[2]))
            self.list2.clear()
            self.find.set("Pretrazi...")


    def edit(self):
        if len(self.display.selection()) == 1:
            self.change = tk.Toplevel()
            self.change.title("Izmena")
            self.contact_for_change = self.display.selection()[0]
            self.change_list = list(self.display.item(self.contact_for_change).values())[2]
            self.previous_name = self.change_list[0]
            self.previous_lastname = self.change_list[1]
            self.previous_phone_number = str(self.change_list[2])
            self.new_name = StringVar()
            self.new_lastname = StringVar()
            self.new_phone_number = StringVar()

            tk.Label(self.change, text="Unesite novo ime:").grid(row=0, column=0)
            tk.Label(self.change, text="Unesite novo prezime:").grid(row=1, column=0)
            tk.Label(self.change, text="Unesite nov broj telefona:").grid(row=2, column=0)
            tk.Button(self.change, text="Izmeni", command=self.change_contact).grid(row=3, column=0, columnspan=2, sticky=W + E)
            self.newNameEntry = tk.Entry(self.change, textvariable=self.new_name)
            self.newLastnameEntry = tk.Entry(self.change, textvariable=self.new_lastname)
            self.newPhoneNumberEntry = tk.Entry(self.change, textvariable=self.new_phone_number)
            self.newNameEntry.grid(row=0, column=1)
            self.newLastnameEntry.grid(row=1, column=1)
            self.newPhoneNumberEntry.grid(row=2, column=1)

            self.newNameEntry.insert(0, self.change_list[0])
            self.newLastnameEntry.insert(0, self.change_list[1])
            self.newPhoneNumberEntry.insert(0, str(self.change_list[2]))

        elif len(self.display.selection()) < 1:
            messagebox.showwarning("Obavestenje","Niste selektovali kontakt koji zelite da izmenite.")
        else:
            messagebox.showwarning("Obavestenje","Ne mozete izmeniti vise kontakata istovremeno. Selektujte samo jedan kontakt.")


    def change_contact(self):
        if self.validation(self.new_name, self.new_lastname, self.new_phone_number, self.newNameEntry, self.newLastnameEntry, self.newPhoneNumberEntry) == True:
            id = 0
            for contact in self.list1:
                if contact == (self.previous_name, self.previous_lastname, self.previous_phone_number):
                    self.list1.pop(id)
                    self.list1.insert(id, (self.new_name.get(), self.new_lastname.get(), self.new_phone_number.get()))
                    break
                else:
                    id += 1
            self.display.delete(*self.display.get_children())
            for contact in self.list1:
                self.display.insert("", "end", values=((contact[0]), (contact[1]), (contact[2])))

            self.change.destroy()

            question = messagebox.askquestion("Pitanje", "Da li zelite da izmenite ovaj kontakt i u bazi?", icon="question")
            if question == "yes":
                try:
                    contacts = load_workbook(filename="ContactsDatabase.xlsx")
                    contacts.active = 0
                    Sheet = contacts.active
                    change = False
                    number = 2
                    for value in Sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
                        if value[0] == self.previous_name and value[1] == self.previous_lastname and str(value[2]) == str(self.previous_phone_number):
                            Sheet["A" + str(number)] = self.new_name.get()
                            Sheet["B" + str(number)] = self.new_lastname.get()
                            Sheet["C" + str(number)] = self.new_phone_number.get()
                            contacts.save(filename="ContactsDatabase.xlsx")
                            change = True
                        else:
                            number += 1

                    if change == True:
                        messagebox.showinfo("Obavestenje", "Zeljena promena kontakta je izvrsena u bazi kontakata.")
                    else:
                        messagebox.showinfo("Obavestenje", "Kontakt koji zelite promeniti se ne nalazi u bazi kontakata.")

                except:
                    messagebox.showinfo("Obavestenje","Ne postoji baza sa kontaktima.")
            else:
                messagebox.showinfo("Obavestenje","Kontakt je promenjen na listi kontakata, ali ne i u bazi.")


    def importDatabase(self):
        try:
            contacts = load_workbook(filename="ContactsDatabase.xlsx")
            contacts.active = 0
            Sheet = contacts.active

            for value in Sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
                if value[0]!=None and value[1]!=None and value[2]!=None:
                    self.display.insert("", "end", values=((value[0]), (value[1]), (value[2])))
                    self.list1.append((value[0], value[1], value[2]))
                else:
                    pass
            if self.list1 == []:
                messagebox.showwarning("Obavestenje","Baza kontakata je prazna.")
        except:
            messagebox.showinfo("Obavestenje","Ne postoji baza sa kontaktima.")


    def deleteDatabase(self):
        warning = messagebox.askquestion("Upozorenje","Upozorenje! Bice obrisana baza i svi kontakti u njoj. Da li ste sigurni da zelite da nastavite?",icon="warning")
        if warning == "yes":
            try:
                filePath = os.path.join(os.getcwd(), "ContactsDatabase.xlsx")
                os.remove(filePath)
                messagebox.showinfo("Obavestenje", "Baza sa kontaktima je uspesno obrisana.")
            except:
                messagebox.showinfo("Obavestenje", "Ne postoji baza sa kontaktima.")
        else:
            pass


start = PhoneBook()